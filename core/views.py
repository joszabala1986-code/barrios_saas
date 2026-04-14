# IMPORTAMOS PARA ANALISIS
import zipfile
import os
from io import BytesIO

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.views.decorators.cache import never_cache
from django.core.mail import send_mail
from django.db import models as django_models
from django.utils import timezone
from django.utils.timezone import now
from django.http import HttpResponse
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db.models import Sum
from django.db.models.functions import TruncMonth
from django.views.decorators.csrf import csrf_protect

from datetime import timedelta, datetime, date
import openpyxl
import json

from .models import (
    Comunicado, Usuario, Deuda, Pago, Propiedad, Barrio,
    Factura, ContratoBarrio, Seguridad, Movimiento,
    SuscripcionBarrio, PagoAdelantado
)
from .utils import (
    subir_comprobante as subir_a_supabase,
    subir_factura as subir_factura_a_supabase,
    subir_contrato as subir_contrato_a_supabase,
    subir_comprobante_adelanto,
    url_firmada_contrato,
)
from .decorators import admin_barrio_required

MESES_NOMBRES = [
    'Enero','Febrero','Marzo','Abril','Mayo','Junio',
    'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre'
]


# ─────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────

def validar_archivo(archivo):
    extensiones_permitidas = ['.pdf', '.jpg', '.jpeg', '.png']
    if not any(archivo.name.lower().endswith(ext) for ext in extensiones_permitidas):
        raise ValidationError("Formato de archivo no permitido")
    if archivo.size > 5 * 1024 * 1024:
        raise ValidationError("El archivo es demasiado grande")


def barrio_usuario(request):
    return request.user.barrio


# ─────────────────────────────────────────
# EXPORTAR COMPROBANTES
# ─────────────────────────────────────────

def descargar_comprobantes(request):
    usuario = request.user
    pagos = Pago.objects.filter(
        deuda__propiedad__propietario=usuario,
        estado='aprobado'
    )
    return render(request, "descargar_comprobantes.html", {"pagos": pagos})


# ─────────────────────────────────────────
# HISTORIAL PROPIETARIO
# ─────────────────────────────────────────

@login_required
def exportar_historial_propietario(request):
    usuario = request.user
    deudas = Deuda.objects.filter(
        propiedad__propietario=usuario
    ).select_related('propiedad', 'propiedad__barrio').order_by('-vencimiento')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial deudas"
    ws.append(["Barrio", "Lote", "Concepto", "Monto", "Estado", "Vencimiento"])

    for deuda in deudas:
        ws.append([
            deuda.propiedad.barrio.nombre,
            deuda.propiedad.numero_lote,
            deuda.concepto,
            float(deuda.monto),
            deuda.estado,
            deuda.vencimiento.strftime("%d/%m/%Y")
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=historial_deudas_propietario.xlsx'
    wb.save(response)
    return response


# ─────────────────────────────────────────
# HISTORIAL LOTES (ADMIN)
# ─────────────────────────────────────────

@admin_barrio_required
def exportar_historial_lotes(request):
    barrio = request.user.barrio
    fecha_limite = timezone.now() - timedelta(days=180)

    deudas = Deuda.objects.filter(
        propiedad__barrio=barrio,
        vencimiento__gte=fecha_limite
    ).select_related('propiedad', 'propiedad__propietario')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial Lotes"
    ws.append(["Lote", "Propietario", "Concepto", "Monto", "Estado", "Vencimiento"])

    for deuda in deudas:
        ws.append([
            deuda.propiedad.numero_lote,
            f"{deuda.propiedad.propietario.first_name} {deuda.propiedad.propietario.last_name}",
            deuda.concepto,
            float(deuda.monto),
            deuda.estado,
            deuda.vencimiento.strftime("%d/%m/%Y")
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=historial_lotes_ultimos6meses.xlsx'
    wb.save(response)
    return response


# ─────────────────────────────────────────
# LOGIN
# ─────────────────────────────────────────

@never_cache
@csrf_protect
def login_view(request):
    if request.user.is_authenticated:
        if request.user.rol == 'superadmin':
            return redirect('panel_superadmin')
        elif request.user.rol == 'admin':
            return redirect('dashboard')
        else:
            return redirect('mis_deudas')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            if user.rol == 'superadmin':
                return redirect('panel_superadmin')
            elif user.rol == 'admin':
                return redirect('dashboard')
            else:
                return redirect('mis_deudas')
        else:
            return render(request, 'login.html', {'error': 'Usuario o contraseña incorrectos'})

    return render(request, 'login.html')


# ─────────────────────────────────────────
# MIS DEUDAS (PROPIETARIO)
# ─────────────────────────────────────────

@login_required
@never_cache
def mis_deudas(request):
    barrio = request.user.barrio
    anio_actual = datetime.now().year

    deudas_pendientes = Deuda.objects.filter(
        propiedad__propietario=request.user,
        estado='pendiente'
    )

    deudas_pagadas = Deuda.objects.filter(
        propiedad__propietario=request.user,
        estado='pagada'
    )

    cantidad_deudas = deudas_pendientes.count()

    propiedad = Propiedad.objects.filter(
        propietario=request.user
    ).first()

    # COMUNICADO con fecha límite
    hoy = now().date()
    comunicado = Comunicado.objects.filter(
        barrio=barrio,
    ).filter(
        django_models.Q(fecha_limite__isnull=True, fecha__gte=timezone.now() - timedelta(days=7)) |
        django_models.Q(fecha_limite__gte=hoy)
    ).order_by("-fecha").first()

    # DEUDAS VENCIDAS
    for deuda in deudas_pendientes:
        deuda.vencida = deuda.vencimiento < hoy

    # GASTOS DEL BARRIO
    mes_actual = datetime.now().month
    facturas = Factura.objects.filter(barrio=barrio).order_by('-fecha')[:10]
    gastos = []
    gastos_mes = 0
    for f in facturas:
        gastos.append((f.descripcion, f.monto, f.fecha))
        if f.fecha and f.fecha.month == mes_actual:
            gastos_mes += float(f.monto)

    # MESES ADELANTADOS DEL PROPIETARIO
    meses_pagados = []
    comprobante_adelanto = None

    if propiedad:
        adelantos = PagoAdelantado.objects.filter(
            propiedad=propiedad,
            anio=anio_actual,
            estado='aprobado'
        )
        for a in adelantos:
            meses_pagados.extend(a.meses)
            if not comprobante_adelanto and a.comprobante:
                comprobante_adelanto = a.comprobante

    meses_pagados = list(set(meses_pagados))

    calendario_meses = [
        {
            'numero': i + 1,
            'nombre': MESES_NOMBRES[i],
            'pagado': (i + 1) in meses_pagados
        }
        for i in range(12)
    ]

    return render(request, 'mis_deudas.html', {
        'deudas': deudas_pendientes,
        'deudas_pagadas': deudas_pagadas,
        'barrio': barrio,
        'propiedad': propiedad,
        'comunicado': comunicado,
        'gastos_mes': gastos_mes,
        'gastos_barrio': gastos,
        'cantidad_deudas': cantidad_deudas,
        'meses_adelantados': meses_pagados,
        'calendario_meses': calendario_meses,
        'comprobante_adelanto': comprobante_adelanto,
        'anio_actual': anio_actual,
    })


# ─────────────────────────────────────────
# SUBIR COMPROBANTE
# ─────────────────────────────────────────

@login_required
@never_cache
def subir_comprobante(request, deuda_id):
    deuda = get_object_or_404(
        Deuda,
        id=deuda_id,
        propiedad__propietario=request.user
    )

    if deuda.estado != 'pendiente':
        return redirect('mis_deudas')

    if request.method == "POST":
        archivo = request.FILES.get('comprobante')

        if not archivo:
            return HttpResponse("Debe subir un archivo")

        try:
            validar_archivo(archivo)
        except ValidationError as e:
            return HttpResponse(str(e))

        url = subir_a_supabase(archivo)

        Pago.objects.create(
            deuda=deuda,
            comprobante=url,
            estado='pendiente_validacion'
        )

        return render(request, 'comprobante_enviado.html')

    return render(request, 'subir_comprobante.html', {'deuda': deuda})


# ─────────────────────────────────────────
# PAGOS PENDIENTES (ADMIN)
# ─────────────────────────────────────────

@admin_barrio_required
@never_cache
def pagos_pendientes(request):
    barrio = request.user.barrio
    pagos = Pago.objects.filter(
        deuda__propiedad__barrio=barrio,
        estado='pendiente_validacion'
    )
    return render(request, 'pagos_pendientes.html', {'pagos': pagos})


@admin_barrio_required
@never_cache
def aprobar_pago(request, pago_id):
    pago = get_object_or_404(
        Pago,
        id=pago_id,
        deuda__propiedad__barrio=request.user.barrio
    )
    pago.estado = 'aprobado'
    pago.save()
    deuda = pago.deuda
    deuda.estado = 'pagada'
    deuda.save()
    return redirect('pagos_pendientes')


@admin_barrio_required
@never_cache
def rechazar_pago(request, pago_id):
    pago = get_object_or_404(Pago, id=pago_id)
    pago.estado = 'rechazado'
    pago.save()
    return redirect('pagos_pendientes')


# ─────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────

@admin_barrio_required
@never_cache
def dashboard(request):
    barrio = barrio_usuario(request)

    mes_param = request.GET.get('mes')
    anio_param = request.GET.get('anio')
    hoy = datetime.now()
    mes_filtro = int(mes_param) if mes_param else hoy.month
    anio_filtro = int(anio_param) if anio_param else hoy.year

    total_recaudado = Pago.objects.filter(
        deuda__propiedad__barrio=barrio,
        estado='aprobado'
    ).aggregate(Sum('deuda__monto'))['deuda__monto__sum'] or 0

    pagos_pendientes_count = Pago.objects.filter(
        deuda__propiedad__barrio=barrio,
        estado='pendiente_validacion'
    ).count()

    deudas_vencidas = Deuda.objects.filter(
        propiedad__barrio=barrio,
        vencimiento__lt=timezone.now(),
        estado='pendiente'
    ).count()

    recaudacion_mensual = (
        Pago.objects
        .filter(deuda__propiedad__barrio=barrio, estado='aprobado')
        .annotate(mes=TruncMonth('fecha_subida'))
        .values('mes')
        .annotate(total=Sum('deuda__monto'))
        .order_by('mes')
    )

    meses = []
    totales = []
    for r in recaudacion_mensual:
        if r['mes'] and r['total']:
            meses.append(r['mes'].strftime("%m/%Y"))
            totales.append(float(r['total']))

    pendientes = Deuda.objects.filter(
        propiedad__barrio=barrio, estado='pendiente'
    ).select_related('propiedad', 'propiedad__propietario')

    vencidas = Deuda.objects.filter(
        propiedad__barrio=barrio,
        estado='pendiente',
        vencimiento__lt=timezone.now()
    ).select_related('propiedad', 'propiedad__propietario')

    esperando = Pago.objects.filter(
        deuda__propiedad__barrio=barrio,
        estado='pendiente_validacion'
    ).select_related('deuda', 'deuda__propiedad', 'deuda__propiedad__propietario')

    pagadas = Deuda.objects.filter(
        propiedad__barrio=barrio, estado='pagada'
    ).select_related('propiedad', 'propiedad__propietario')

    total_deudas = Deuda.objects.filter(propiedad__barrio=barrio).count()
    deudas_pagadas = Deuda.objects.filter(propiedad__barrio=barrio, estado='pagada').count()
    deudas_pendientes = Deuda.objects.filter(propiedad__barrio=barrio, estado='pendiente').count()

    morosidad = 0
    if total_deudas > 0:
        morosidad = round((deudas_pendientes / total_deudas) * 100, 2)

    facturas_barrio = Factura.objects.filter(barrio=barrio)
    total_gastos = facturas_barrio.aggregate(Sum('monto'))['monto__sum'] or 0
    gastos_mes = facturas_barrio.filter(
        fecha__month=mes_filtro,
        fecha__year=anio_filtro
    ).aggregate(Sum('monto'))['monto__sum'] or 0
    caja_disponible = total_recaudado - total_gastos

    pagos_mes = Pago.objects.filter(
        deuda__propiedad__barrio=barrio,
        fecha_subida__month=mes_filtro,
        fecha_subida__year=anio_filtro,
        estado='aprobado'
    ).select_related('deuda', 'deuda__propiedad', 'deuda__propiedad__propietario')

    facturas_mes = facturas_barrio.filter(
        fecha__month=mes_filtro,
        fecha__year=anio_filtro
    )

    return render(request, 'dashboard.html', {
        'barrio': barrio,
        'total_recaudado': total_recaudado,
        'pagos_pendientes': pagos_pendientes_count,
        'deudas_vencidas': deudas_vencidas,
        'lotes_pendientes': pendientes,
        'lotes_vencidos': vencidas,
        'lotes_validacion': esperando,
        'lotes_pagados': pagadas,
        'meses': json.dumps(meses),
        'totales': json.dumps(totales),
        'deudas_pagadas': deudas_pagadas,
        'deudas_pendientes': deudas_pendientes,
        'morosidad': morosidad,
        'gastos_mes': gastos_mes,
        'total_gastos': total_gastos,
        'caja_disponible': caja_disponible,
        'mes_filtro': mes_filtro,
        'anio_filtro': anio_filtro,
        'pagos_mes': pagos_mes,
        'facturas_mes': facturas_mes,
    })


# ─────────────────────────────────────────
# LOGOUT
# ─────────────────────────────────────────

def logout_view(request):
    logout(request)
    request.session.flush()
    return redirect('/login/')


# ─────────────────────────────────────────
# MOROSOS
# ─────────────────────────────────────────

@admin_barrio_required
def morosos(request):
    hoy = timezone.now().date()
    barrio = barrio_usuario(request)
    deudas_vencidas = Deuda.objects.filter(
        propiedad__barrio=barrio,
        estado='pendiente',
        vencimiento__lt=hoy
    ).select_related('propiedad', 'propiedad__propietario', 'propiedad__barrio')
    return render(request, 'morosos.html', {'deudas': deudas_vencidas})


# ─────────────────────────────────────────
# FACTURAS
# ─────────────────────────────────────────

@admin_barrio_required
def eliminar_factura(request, factura_id):
    factura = Factura.objects.get(id=factura_id, barrio=request.user.barrio)
    factura.delete()
    return redirect("facturas")


@admin_barrio_required
def exportar_morosos_excel(request):
    hoy = timezone.now().date()
    deudas_vencidas = Deuda.objects.filter(
        estado='pendiente',
        vencimiento__lt=hoy
    ).select_related('propiedad', 'propiedad__propietario', 'propiedad__barrio')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Morosos"
    ws.append(["Propietario", "Lote", "Barrio", "Concepto", "Monto", "Vencimiento"])

    for deuda in deudas_vencidas:
        ws.append([
            f"{deuda.propiedad.propietario.first_name} {deuda.propiedad.propietario.last_name}",
            deuda.propiedad.numero_lote,
            deuda.propiedad.barrio.nombre,
            deuda.concepto,
            float(deuda.monto),
            deuda.vencimiento.strftime("%d/%m/%Y")
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=morosos.xlsx'
    wb.save(response)
    return response


@admin_barrio_required
def facturas(request):
    barrio = request.user.barrio
    facturas_qs = Factura.objects.filter(barrio=barrio).order_by("-fecha")
    total_gastos = facturas_qs.aggregate(Sum("monto"))["monto__sum"] or 0

    mes_actual = datetime.now().month
    gastos_mes = 0
    meses_lista = [0] * 12

    for f in facturas_qs:
        if not f.fecha:
            continue
        mes = f.fecha.month
        meses_lista[mes - 1] += float(f.monto)
        if mes == mes_actual:
            gastos_mes += float(f.monto)

    total_recaudado = Pago.objects.filter(
        deuda__propiedad__barrio=barrio,
        estado='aprobado'
    ).aggregate(Sum('deuda__monto'))['deuda__monto__sum'] or 0

    caja_disponible = total_recaudado - total_gastos

    return render(request, "facturas.html", {
        "facturas": facturas_qs,
        "gastos_mes": gastos_mes,
        "caja_disponible": caja_disponible,
        "gastos_meses": meses_lista
    })


@admin_barrio_required
def guardar_factura(request):
    if request.method == "POST":
        from decimal import Decimal, InvalidOperation
        descripcion = request.POST.get("descripcion")
        monto_input = request.POST.get("monto")
        archivo = request.FILES.get("archivo")
        barrio = request.user.barrio

        try:
            monto_input = monto_input.replace(".", "").replace(",", ".")
            monto = Decimal(monto_input)
            if monto <= 0:
                raise ValueError
        except (InvalidOperation, ValueError):
            return HttpResponse("Monto inválido")

        if not descripcion:
            return HttpResponse("Debe ingresar una descripción")

        url_archivo = None
        if archivo:
            try:
                validar_archivo(archivo)
            except ValidationError as e:
                return HttpResponse(str(e))
            url_archivo = subir_factura_a_supabase(archivo)

        Factura.objects.create(
            barrio=barrio,
            descripcion=descripcion,
            monto=monto,
            archivo=url_archivo
        )

    return redirect("facturas")


@admin_barrio_required
def editar_factura(request, factura_id):
    factura = Factura.objects.get(id=factura_id, barrio=request.user.barrio)

    if request.method == "POST":
        factura.descripcion = request.POST.get("descripcion")
        monto = request.POST.get("monto").replace(".", "").replace(",", ".")
        factura.monto = monto

        archivo = request.FILES.get("archivo")
        if archivo:
            factura.archivo = subir_factura_a_supabase(archivo)

        factura.save()
        return redirect("facturas")

    return render(request, "editar_factura.html", {"factura": factura})


# ─────────────────────────────────────────
# LOTES Y DEUDAS
# ─────────────────────────────────────────

@admin_barrio_required
def generar_deuda_masiva(request):
    barrio = barrio_usuario(request)

    if request.method == "POST":
        from decimal import Decimal, InvalidOperation

        try:
            concepto = request.POST.get("concepto")
            monto_input = request.POST.get("monto")
            vencimiento = request.POST.get("vencimiento")
            mes_seleccionado = request.POST.get("mes_seleccionado")
            monto = Decimal(monto_input)
            if not concepto or monto <= 0 or not vencimiento or not mes_seleccionado:
                raise ValueError
        except (InvalidOperation, ValueError, TypeError):
            return HttpResponse("Datos inválidos — asegurate de seleccionar un mes")

        mes_expensa = int(mes_seleccionado)
        anio_expensa = datetime.now().year
        propiedades = Propiedad.objects.filter(barrio=barrio)
        generadas = 0
        salteadas = 0

        for propiedad in propiedades:
            ya_pago = PagoAdelantado.objects.filter(
                propiedad=propiedad,
                anio=anio_expensa,
                estado='aprobado',
                meses__contains=mes_expensa
            ).exists()

            if ya_pago:
                Deuda.objects.create(
                    propiedad=propiedad,
                    concepto=concepto,
                    descripcion=f"{concepto} — abonado en pago adelantado",
                    monto=monto,
                    vencimiento=vencimiento,
                    estado='pagada'
                )
                salteadas += 1
            else:
                Deuda.objects.create(
                    propiedad=propiedad,
                    concepto=concepto,
                    descripcion=concepto,
                    monto=monto,
                    vencimiento=vencimiento,
                    estado='pendiente'
                )
                generadas += 1

        messages.success(request,
            f"✅ Expensas generadas: {generadas} pendientes, {salteadas} ya cubiertas por pago adelantado")
        return redirect("dashboard")

    return render(request, "generar_deuda.html", {"barrio": barrio})


@admin_barrio_required
def crear_propietario(request):
    barrio = barrio_usuario(request)

    if request.method == "POST":
        username = request.POST.get("username")
        nombre = request.POST.get("nombre")
        apellido = request.POST.get("apellido")
        email = request.POST.get("email")
        password = request.POST.get("password")
        lote = request.POST.get("lote")
        celular = request.POST.get("celular")

        if Usuario.objects.filter(username=username).exists():
            messages.error(request, "Ese usuario ya existe")
            return redirect("crear_propietario")

        if Propiedad.objects.filter(numero_lote=lote, barrio=barrio).exists():
            messages.error(request, "Ese lote ya existe en el barrio")
            return redirect("crear_propietario")

        usuario = Usuario.objects.create_user(
            username=username,
            password=password,
            first_name=nombre,
            last_name=apellido,
            email=email,
            celular=celular,
            rol="propietario",
            barrio=barrio
        )

        Propiedad.objects.create(
            barrio=barrio,
            numero_lote=lote,
            propietario=usuario
        )

        messages.success(request, "Propietario y lote creados correctamente")
        return redirect("lotes")

    return render(request, "crear_propietario.html")


@admin_barrio_required
def lotes_barrio(request):
    barrio = barrio_usuario(request)
    propiedades = Propiedad.objects.filter(barrio=barrio).select_related('propietario')
    anio_actual = datetime.now().year

    for propiedad in propiedades:
        adelantos = PagoAdelantado.objects.filter(
            propiedad=propiedad,
            anio=anio_actual,
            estado='aprobado'
        )
        meses_pagados = []
        for a in adelantos:
            meses_pagados.extend(a.meses)
        propiedad.meses_pagados_json = json.dumps(list(set(meses_pagados)))

    return render(request, 'lotes.html', {
        'propiedades': propiedades,
        'barrio': barrio,
        'anio_actual': anio_actual,
    })


@admin_barrio_required
def crear_deuda_lote(request, lote_id):
    propiedad = Propiedad.objects.get(id=lote_id)

    if request.method == "POST":
        concepto = request.POST.get("concepto")
        monto = request.POST.get("monto")
        vencimiento = request.POST.get("vencimiento")

        Deuda.objects.create(
            propiedad=propiedad,
            concepto=concepto,
            descripcion=concepto,
            monto=monto,
            vencimiento=vencimiento,
            estado="pendiente"
        )

        messages.success(request, "Deuda creada correctamente")
        return redirect("lotes")

    return render(request, "crear_deuda_lote.html", {"propiedad": propiedad})


@admin_barrio_required
def ficha_lote(request, lote_id):
    propiedad = get_object_or_404(
        Propiedad,
        id=lote_id,
        barrio=request.user.barrio
    )

    deudas = Deuda.objects.filter(propiedad=propiedad, estado="pendiente")
    pagos = Pago.objects.filter(deuda__propiedad=propiedad, estado="aprobado")

    return render(request, "ficha_lote.html", {
        "propiedad": propiedad,
        "deudas": deudas,
        "pagos": pagos
    })


@admin_barrio_required
def eliminar_deuda(request, deuda_id):
    deuda = get_object_or_404(
        Deuda,
        id=deuda_id,
        propiedad__barrio=request.user.barrio
    )
    deuda.delete()
    return redirect("lotes")


@admin_barrio_required
def pago_efectivo(request, deuda_id):
    deuda = get_object_or_404(
        Deuda,
        id=deuda_id,
        propiedad__barrio=request.user.barrio
    )
    deuda.estado = "pagada"
    deuda.save()

    Pago.objects.create(deuda=deuda, estado="aprobado")
    return redirect("ficha_lote", lote_id=deuda.propiedad.id)


# ─────────────────────────────────────────
# PAGO ADELANTADO
# ─────────────────────────────────────────

@admin_barrio_required
def pago_adelantado(request):
    if request.method != 'POST':
        return redirect('lotes')

    from decimal import Decimal

    propiedad_id = request.POST.get('propiedad_id')
    monto_input = request.POST.get('monto', '0').replace('.', '').replace(',', '.')
    meses_seleccionados = request.POST.getlist('meses')
    archivo = request.FILES.get('comprobante')
    anio_actual = datetime.now().year

    if not meses_seleccionados:
        return HttpResponse("Debe seleccionar al menos un mes")

    try:
        monto = Decimal(monto_input)
        if monto <= 0:
            raise ValueError
    except Exception:
        return HttpResponse("Monto inválido")

    propiedad = get_object_or_404(
        Propiedad,
        id=propiedad_id,
        barrio=request.user.barrio
    )

    meses_int = [int(m) for m in meses_seleccionados]
    meses_str = ", ".join([MESES_NOMBRES[m - 1] for m in meses_int])

    url_comprobante = None
    if archivo:
        try:
            validar_archivo(archivo)
            url_comprobante = subir_comprobante_adelanto(archivo)
        except ValidationError as e:
            return HttpResponse(str(e))

    deuda = Deuda.objects.create(
        propiedad=propiedad,
        concepto=f"Expensa adelantada {anio_actual}",
        descripcion=f"Pago adelantado meses: {meses_str}",
        monto=monto,
        vencimiento=f"{anio_actual}-12-31",
        estado='pagada'
    )

    Pago.objects.create(
        deuda=deuda,
        comprobante=url_comprobante or '',
        estado='aprobado'
    )

    PagoAdelantado.objects.create(
        propiedad=propiedad,
        meses=meses_int,
        anio=anio_actual,
        monto=monto,
        comprobante=url_comprobante,
        estado='aprobado'
    )

    messages.success(request, f"Pago adelantado registrado: meses {meses_str}")
    return redirect('lotes')


# ─────────────────────────────────────────
# CONTRATO
# ─────────────────────────────────────────

@login_required
def ver_contrato(request):
    barrio = request.user.barrio
    
    # Buscar el contrato más reciente del barrio
    contrato = ContratoBarrio.objects.filter(
        barrio=barrio
    ).order_by('-fecha_subida').first()

    url_temporal = None
    if contrato and contrato.archivo:
        # Generar URL firmada usando el path guardado
        from .utils import url_firmada_contrato
        url_temporal = url_firmada_contrato(contrato.archivo)

    return render(request, 'ver_contrato.html', {
        'contrato': contrato,
        'url_contrato': url_temporal  # ← Esta es la clave!
    })


@login_required
def subir_contrato(request):
    # Solo admin o staff pueden subir contrato del barrio
    if not request.user.is_staff and request.user.rol != 'admin':
        messages.error(request, "No tenés permiso para subir contratos")
        return redirect('mis_deudas')

    barrio = request.user.barrio
    if not barrio:
        messages.error(request, "No tenés un barrio asignado")
        return redirect('dashboard')

    if request.method == 'POST':
        archivo = request.FILES.get('contrato')
        
        if not archivo:
            messages.error(request, "❌ Seleccioná un archivo")
            return redirect('subir_contrato')
        
        try:
            # Validar archivo
            validar_archivo(archivo)
            
            # Subir a Supabase (devuelve el PATH, no la URL)
            from .utils import subir_contrato as subir_contrato_a_supabase
            path = subir_contrato_a_supabase(archivo)
            
            # Guardar en el modelo ContratoBarrio (el path del archivo)
            ContratoBarrio.objects.create(
                barrio=barrio,
                archivo=path  # Guarda el path, no la URL
            )
            
            messages.success(request, "✅ Contrato del barrio subido exitosamente")
            return redirect('dashboard')
            
        except ValidationError as e:
            messages.error(request, f"❌ Error: {str(e)}")
            return redirect('subir_contrato')
        except Exception as e:
            messages.error(request, f"❌ Error al subir: {str(e)}")
            return redirect('subir_contrato')

    return render(request, 'subir_contrato.html', {'barrio': barrio})


# ─────────────────────────────────────────
# COMUNICADOS
# ─────────────────────────────────────────

from django.contrib import messages  # Asegúrate de tener esto importado al inicio

@admin_barrio_required
def crear_comunicado(request):
    barrio = barrio_usuario(request)

    if request.method == "POST":
        mensaje = request.POST.get("mensaje")
        fecha_limite = request.POST.get("fecha_limite")
        
        # Convertir string vacío a None
        if not fecha_limite or fecha_limite == "":
            fecha_limite = None
        
        # Crear el comunicado
        Comunicado.objects.create(
            barrio=barrio,
            mensaje=mensaje,
            fecha_limite=fecha_limite
        )
        
        # Mostrar mensaje de éxito
        messages.success(request, "✅ ¡Comunicado enviado exitosamente!")
        
        # Redirigir a la lista de lotes
        return redirect("lotes")  # Este nombre existe en tu urls.py

    # GET request - mostrar formulario
    hoy = timezone.now().date().isoformat()
    return render(request, "crear_comunicado.html", {
        "barrio": barrio,
        "hoy": hoy
    })
# ─────────────────────────────────────────
# CAMBIAR PROPIETARIO
# ─────────────────────────────────────────

@admin_barrio_required
def cambiar_propietario(request, lote_id):
    propiedad = Propiedad.objects.select_related("propietario").get(id=lote_id)
    usuario = propiedad.propietario

    if request.method == "POST":
        usuario.username = request.POST.get("username")
        usuario.first_name = request.POST.get("nombre")
        usuario.last_name = request.POST.get("apellido")
        usuario.email = request.POST.get("email")
        usuario.celular = request.POST.get("celular")

        password = request.POST.get("password")
        if password:
            usuario.set_password(password)

        usuario.save()
        return redirect("ficha_lote", lote_id=lote_id)

    return render(request, "cambiar_propietario.html", {
        "propiedad": propiedad,
        "usuario": usuario
    })


# ─────────────────────────────────────────
# PANEL PLANTA / SUPERADMIN
# ─────────────────────────────────────────

@admin_barrio_required
def panel_planta(request):
    barrios = Barrio.objects.all()
    usuarios = Usuario.objects.count()
    return render(request, "panel_planta.html", {
        "barrios": barrios,
        "usuarios": usuarios
    })


@login_required
def panel_superadmin(request):
    if not request.user.rol == 'superadmin':
        return redirect('login')

    barrios = Barrio.objects.all()
    usuarios = Usuario.objects.count()
    today = date.today()

    for b in barrios:

            b.total_usuarios = Propiedad.objects.filter(
                barrio=b
            ).count()

            if hasattr(b, 'suscripcionbarrio'):
                b.dias_restantes = (b.suscripcionbarrio.fecha_vencimiento - today).days
            else:
                b.dias_restantes = None

            b.total_recaudado = Pago.objects.filter(
                deuda__propiedad__barrio=b,
                estado='aprobado'
            ).aggregate(Sum('deuda__monto'))['deuda__monto__sum'] or 0

    return render(request, 'panel_superadmin.html', {
        'barrios': barrios,
        'usuarios': usuarios,
        'today': today
    })


@login_required
def crear_barrio(request):
    if request.user.rol != 'superadmin':
        return redirect('login')

    if request.method == "POST":
        nombre = request.POST.get("nombre")
        cbu = request.POST.get("cbu")
        telefono = request.POST.get("telefono")
        titular_cuenta = request.POST.get("titular_cuenta", "Asociación Vecinal")
        cuit = request.POST.get("cuit", "30-12345678-9")
        banco = request.POST.get("banco", "Banco Nación")
        plan = request.POST.get("plan", "mensual")
        precio_suscripcion = request.POST.get("precio_suscripcion", 0)
        admin_usuario = request.POST.get("admin_usuario")
        admin_password = request.POST.get("admin_password")

        if Usuario.objects.filter(username=admin_usuario).exists():
            return HttpResponse("El usuario admin ya existe")

        barrio = Barrio.objects.create(
            nombre=nombre,
            cbu=cbu,
            telefono=telefono,
            titular_cuenta=titular_cuenta,
            cuit=cuit,
            banco=banco,
            plan=plan,
            precio_suscripcion=precio_suscripcion,
            fecha_inicio_suscripcion=date.today()
        )

        Usuario.objects.create_user(
            username=admin_usuario,
            password=admin_password,
            rol="admin",
            barrio=barrio,
            is_staff=True
        )

        SuscripcionBarrio.objects.create(
            barrio=barrio,
            plan=plan,
            precio=precio_suscripcion,
            fecha_vencimiento=date.today() + timedelta(days=30),
            activo=True
        )

        return redirect('panel_superadmin')

    return render(request, "crear_barrio.html")


@login_required
def eliminar_barrio(request, barrio_id):
    if request.user.rol != 'superadmin':
        return redirect('login')
    barrio = get_object_or_404(Barrio, id=barrio_id)
    barrio.delete()
    return redirect('panel_superadmin')


@login_required
def editar_barrio(request, barrio_id):
    if request.user.rol != 'superadmin':
        return redirect('login')

    barrio = get_object_or_404(Barrio, id=barrio_id)

    if request.method == 'POST':
        barrio.nombre = request.POST.get('nombre')
        barrio.cbu = request.POST.get('cbu')
        barrio.telefono = request.POST.get('telefono')
        barrio.titular_cuenta = request.POST.get('titular_cuenta')
        barrio.cuit = request.POST.get('cuit')
        barrio.banco = request.POST.get('banco')

        fecha = request.POST.get('fecha_vencimiento')
        if fecha:
            barrio.fecha_vencimiento = fecha

        barrio.save()
        return redirect('panel_superadmin')

    return render(request, 'editar_barrio.html', {'barrio': barrio})


@login_required
def marcar_pagado(request, barrio_id):
    if request.user.rol != 'superadmin':
        return redirect('login')

    barrio = get_object_or_404(Barrio, id=barrio_id)

    if barrio.fecha_vencimiento and barrio.fecha_vencimiento >= now().date():
        return redirect('panel_superadmin')

    barrio.marcar_pagado()
    return redirect('panel_superadmin')


# ─────────────────────────────────────────
# EXPORTAR BARRIO EXCEL
# ─────────────────────────────────────────

def exportar_barrio_excel(request):
    barrio = request.user.barrio
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Usuarios"
    ws1.append(["Nombre", "Email"])
    for u in Usuario.objects.filter(barrio=barrio):
        ws1.append([u.nombre_completo, u.email])

    ws2 = wb.create_sheet(title="Pagos")
    ws2.append(["Usuario", "Monto", "Fecha"])
    for p in Pago.objects.filter(deuda__propiedad__barrio=barrio):
        ws2.append([
            p.deuda.propiedad.propietario.nombre_completo,
            p.deuda.monto,
            p.fecha_subida.strftime("%d/%m/%Y %H:%M")
        ])

    ws3 = wb.create_sheet(title="Deudas")
    ws3.append(["Usuario", "Lote", "Concepto", "Monto", "Estado", "Vencimiento", "Fecha de pago"])
    for d in Deuda.objects.filter(propiedad__barrio=barrio).select_related("propiedad", "propiedad__propietario"):
        pago = Pago.objects.filter(deuda=d, estado="aprobado").first()
        fecha_pago = pago.fecha_subida.strftime("%d/%m/%Y %H:%M") if pago else "-"
        ws3.append([
            d.propiedad.propietario.nombre_completo,
            d.propiedad.numero_lote,
            d.concepto,
            float(d.monto),
            d.estado,
            d.vencimiento.strftime("%d/%m/%Y"),
            fecha_pago
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{barrio.nombre}_datos.xlsx"'
    wb.save(response)
    return response


# ─────────────────────────────────────────
# SEGURIDAD
# ─────────────────────────────────────────

@login_required
def panel_seguridad(request, barrio_id):
    if request.user.rol not in ['admin', 'superadmin']:
        return redirect('login')

    if request.user.rol == 'admin' and request.user.barrio.id != barrio_id:
        return HttpResponse("No autorizado")

    barrio = get_object_or_404(Barrio, id=barrio_id)
    ingresos = Movimiento.objects.filter(barrio=barrio, tipo='ingreso').order_by('-fecha_hora')
    egresos = Movimiento.objects.filter(barrio=barrio, tipo='egreso').order_by('-fecha_hora')
    personal = Seguridad.objects.filter(barrio=barrio)

    return render(request, 'panel_seguridad.html', {
        'barrio': barrio,
        'ingresos': ingresos,
        'egresos': egresos,
        'personal': personal
    })


from django.contrib.auth import get_user_model
from .forms import SeguridadForm
User = get_user_model()


def crear_seguridad(request, barrio_id):
    barrio = get_object_or_404(Barrio, id=barrio_id)

    if request.method == 'POST':
        form = SeguridadForm(request.POST, request.FILES)

        if form.is_valid():
            dni = form.cleaned_data['dni']
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']

            if User.objects.filter(username=dni).exists():
                form.add_error('dni', 'DNI ya existente')
            elif User.objects.filter(email=email).exists():
                form.add_error('email', 'Email ya registrado')
            else:
                seguridad = form.save(commit=False)
                seguridad.barrio = barrio
                user = User.objects.create_user(username=dni, email=email, password=password)
                seguridad.usuario = user
                seguridad.save()
                return redirect('panel_seguridad', barrio_id=barrio.id)
    else:
        form = SeguridadForm()

    return render(request, 'crear_seguridad.html', {'form': form, 'barrio': barrio})


def login_seguridad(request):
    error = None
    if request.method == 'POST':
        dni = request.POST.get('dni')
        password = request.POST.get('password')
        user = authenticate(request, username=dni, password=password)
        if user is not None:
            login(request, user)
            return redirect('panel_guardia')
        else:
            error = "DNI o contraseña incorrectos"
    return render(request, 'login_seguridad.html', {'error': error})


@login_required
def panel_guardia(request):
    seguridad = Seguridad.objects.get(usuario=request.user)
    movimientos = Movimiento.objects.filter(
        barrio=seguridad.barrio
    ).order_by('-fecha_hora')[:20]
    return render(request, 'panel_guardia.html', {
        'seguridad': seguridad,
        'movimientos': movimientos
    })


@login_required
def registrar_movimiento(request, tipo):
    seguridad = Seguridad.objects.get(usuario=request.user)

    if request.method == 'POST':
        Movimiento.objects.create(
            barrio=seguridad.barrio,
            tipo=tipo,
            fecha_hora=timezone.now(),
            patente=request.POST.get('patente', '').upper(),
            nombre_apellido=request.POST.get('nombre'),
            dni=request.POST.get('dni', ''),
            observaciones=request.POST.get('observaciones', ''),
            registrado_por=seguridad
        )
        return redirect('panel_guardia')  # ← vuelve al panel con la lista actualizada

    return redirect('panel_guardia')
@admin_barrio_required
def eliminar_seguridad(request, seguridad_id):
    seguridad = get_object_or_404(Seguridad, id=seguridad_id)
    barrio_id = seguridad.barrio.id
    seguridad.usuario.delete()
    seguridad.delete()
    return redirect('panel_seguridad', barrio_id=barrio_id)


@admin_barrio_required
def editar_seguridad(request, seguridad_id):
    seguridad = get_object_or_404(Seguridad, id=seguridad_id)

    if request.method == 'POST':
        form = SeguridadForm(request.POST, request.FILES, instance=seguridad)
        if form.is_valid():
            form.save()
            return redirect('panel_seguridad', barrio_id=seguridad.barrio.id)
    else:
        form = SeguridadForm(instance=seguridad)

    return render(request, 'crear_seguridad.html', {
        'form': form,
        'barrio': seguridad.barrio
    })


def logout_seguridad(request):
    logout(request)
    return redirect('login_seguridad')


@admin_barrio_required
def exportar_excel(request, barrio_id):
    if request.user.barrio.id != barrio_id:
        return HttpResponse("No autorizado")

    movimientos = Movimiento.objects.filter(barrio_id=barrio_id).order_by('fecha_hora')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    ws.append(["Fecha", "Hora", "Tipo", "Nombre y Apellido", "DNI", "Patente", "Observaciones", "Personal Seguridad"])

    for m in movimientos:
        ws.append([
            m.fecha_hora.date(),
            m.fecha_hora.time().strftime('%H:%M:%S'),
            m.tipo.capitalize(),
            m.nombre_apellido,
            m.dni,
            m.patente,
            m.observaciones if m.observaciones else "",
            str(m.registrado_por)
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=movimientos.xlsx'
    wb.save(response)
    return response


# ─────────────────────────────────────────
# DESCARGAR COMPROBANTES
# ─────────────────────────────────────────

def descargar_comprobantes(request):
    usuario = request.user
    pagos = Pago.objects.filter(
        deuda__propiedad__propietario=usuario,
        estado='aprobado'
    )
    return render(request, "descargar_comprobantes.html", {"pagos": pagos})

# ─────────────────────────────────────────
# OFFLINE
# ─────────────────────────────────────────

def offline_view(request):
    return render(request, 'offline.html')


